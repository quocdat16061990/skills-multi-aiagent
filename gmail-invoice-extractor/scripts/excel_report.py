from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DANGEROUS = ("=", "+", "-", "@")
def safe(v):
    if v is None: return ""
    s = str(v)
    return "'" + s if s.startswith(DANGEROUS) else s

def write_report(path: Path, invoices: list[dict], attachments: list[dict], duplicates: list[dict], errors: list[dict], summary: dict) -> None:
    wb = Workbook(); wb.remove(wb.active)
    sheets = [("Summary", [summary]), ("Invoices", invoices), ("Attachments", attachments), ("Duplicates", duplicates), ("Review_Errors", errors)]
    for name, rows in sheets:
        ws = wb.create_sheet(name)
        keys = list(dict.fromkeys(k for r in rows for k in r)) or ["message"]
        ws.append(keys)
        for c in ws[1]: c.font=Font(bold=True, color="FFFFFF"); c.fill=PatternFill("solid", fgColor="1F4E78")
        for r in rows: ws.append([safe(r.get(k, "")) for k in keys])
        ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
        for i, k in enumerate(keys, 1):
            width = max([len(str(k))] + [len(str(r.get(k, ""))) for r in rows])
            ws.column_dimensions[get_column_letter(i)].width=min(width+2, 60)
    path.parent.mkdir(parents=True, exist_ok=True); wb.save(path)
    check = load_workbook(path, read_only=False, data_only=False)
    if check.sheetnames != [x[0] for x in sheets]: raise RuntimeError("workbook sheet verification failed")
    for ws in check.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith(DANGEROUS): raise RuntimeError("unsafe formula-like cell")
    check.close()
