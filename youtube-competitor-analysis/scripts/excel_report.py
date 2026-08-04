"""Seven-sheet Vietnamese Excel report with safety and reopen verification."""
from __future__ import annotations
from collections import Counter
from datetime import datetime,timezone
from openpyxl import Workbook,load_workbook
from openpyxl.chart import BarChart,Reference
from openpyxl.styles import Font,PatternFill,Alignment
from openpyxl.utils import get_column_letter

SHEETS=["Tong_quan","Du_lieu_video","Top_video","Chu_de","Hook_va_tieu_de","De_xuat_hanh_dong","Nguon_va_gioi_han"]
def safe(v):
    if isinstance(v,str) and v[:1] in ("=","+","-","@"):return "'"+v
    return v
def table(ws,headers,rows):
    ws.append(headers)
    for row in rows:ws.append([safe(v) for v in row])
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for c in ws[1]:c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="1F4E78"); c.alignment=Alignment(wrap_text=True)
    for col in range(1,ws.max_column+1):
        ws.column_dimensions[get_column_letter(col)].width=min(50,max(12,max(len(str(ws.cell(r,col).value or "")) for r in range(1,ws.max_row+1))+2))
def write_report(path,channel,rows,stats,cfg,collected):
    wb=Workbook(); wb.remove(wb.active); ws={n:wb.create_sheet(n) for n in SHEETS}
    overview=[]
    for group,m in stats.items():
        for metric in ("count","total","mean","median","min","max"):overview.append([group,metric,m.get(metric)])
    table(ws["Tong_quan"],["Nhóm","Chỉ số","Giá trị"],overview)
    headers=["Kênh","Handle","Subscribers","Video ID","Tiêu đề","URL","Views","Ngày đăng/relative","Ngày chắc chắn","Duration (giây)","Loại","Views/ngày","Chủ đề (coding)","Bằng chứng chủ đề","Hook (coding)","Bằng chứng hook"]
    data=[]
    for r in rows:data.append([channel.get("title"),channel.get("handle"),channel.get("subscribers"),r.get("video_id"),r.get("title"),r.get("url"),r.get("views"),r.get("published"),r.get("published_certain"),r.get("duration"),r.get("type"),r.get("views_per_day"),", ".join(r.get("topics",[])),", ".join(r.get("topic_evidence",[])),", ".join(r.get("hooks",[])),", ".join(r.get("hook_evidence",[]))])
    table(ws["Du_lieu_video"],headers,data)
    for row in range(2,ws["Du_lieu_video"].max_row+1):
        cell=ws["Du_lieu_video"].cell(row,6)
        if isinstance(cell.value,str) and cell.value.startswith("http"):cell.hyperlink=cell.value; cell.style="Hyperlink"
    top=sorted(rows,key=lambda r:r.get("views") if isinstance(r.get("views"),(int,float)) else -1,reverse=True)[:int(cfg.get("recommendations",{}).get("top_n",10))]
    table(ws["Top_video"],["Hạng","Video ID","Tiêu đề","URL","Views","Loại"],[[i,r.get("video_id"),r.get("title"),r.get("url"),r.get("views"),r.get("type")] for i,r in enumerate(top,1)])
    for row in range(2,ws["Top_video"].max_row+1):ws["Top_video"].cell(row,4).hyperlink=ws["Top_video"].cell(row,4).value; ws["Top_video"].cell(row,4).style="Hyperlink"
    tc=Counter(t for r in rows for t in r.get("topics",[])); table(ws["Chu_de"],["Chủ đề (coding)","Số video"],tc.most_common())
    hc=Counter(h for r in rows for h in r.get("hooks",[])); table(ws["Hook_va_tieu_de"],["Hook (coding)","Số video","Ghi chú"],[[k,v,"Nhãn deterministic; xem bằng chứng ở Du_lieu_video"] for k,v in hc.most_common()])
    rec=[]
    if top:rec.append(["Inference","Ưu tiên kiểm thử chủ đề/hook xuất hiện trong nhóm top","Trung bình/đỉnh views của dữ liệu quan sát; không khẳng định nhân quả","Trung bình"])
    if not rec:rec.append(["Inference","Chưa đủ dữ liệu để đề xuất","Không có video hợp lệ","Thấp"])
    table(ws["De_xuat_hanh_dong"],["Loại phát biểu","Đề xuất","Bằng chứng","Độ tin cậy"],rec)
    limits=[["Fact","Nguồn",collected.get("source")],["Fact","URL kênh",channel.get("url")],["Fact","Thời điểm tạo",datetime.now(timezone.utc).isoformat()],["Coding","Phân loại","Từ khóa/regex deterministic theo config; không phải nhãn của tác giả"],["Inference","Đề xuất","Diễn giải tương quan mô tả, không chứng minh nhân quả"],["Giới hạn","Thu thập","Chỉ dữ liệu công khai; không đăng nhập, không vượt chặn; HTML có thể thiếu continuation"],["Giới hạn","Ngày đăng","Views/ngày chỉ tính khi ngày tuyệt đối chắc chắn"],["Giới hạn","Subscribers","Có thể N/A nếu nguồn không công khai/không cung cấp"]]
    limits += [["Cảnh báo","Collector",x] for x in collected.get("warnings",[])]
    table(ws["Nguon_va_gioi_han"],["Phân loại","Mục","Nội dung"],limits)
    if ws["Chu_de"].max_row>1:
        chart=BarChart(); chart.title="Phân bố chủ đề"; chart.add_data(Reference(ws["Chu_de"],min_col=2,min_row=1,max_row=ws["Chu_de"].max_row),titles_from_data=True); chart.set_categories(Reference(ws["Chu_de"],min_col=1,min_row=2,max_row=ws["Chu_de"].max_row)); ws["Tong_quan"].add_chart(chart,"E2")
    wb.save(path); verify_report(path)
def verify_report(path):
    wb=load_workbook(path,data_only=False)
    if wb.sheetnames!=SHEETS:raise RuntimeError(f"Sai cấu trúc sheet: {wb.sheetnames}")
    if wb["Du_lieu_video"].max_column<16:raise RuntimeError("Thiếu cột dữ liệu")
    wb.close(); return True
